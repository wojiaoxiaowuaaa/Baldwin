import openpyxl
import pandas as pd

class ExcelModifier:
    def __init__(self, file_path):
        """
        :param file_path: 待处理的文件.
        初始化方法，用于加载工作簿
        self.file_path 将文件路径作为实例属性保存下来,使得这个路径可以在类的其他方法中被访问和使用.
        """
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(filename=self.file_path)
        self.sheet = self.workbook.active

    def duplicate_first_row(self, output_path):
        """
        将第二行的数据复制到后续行.
        :param output_path: 处理后输出文件路径
        """
        first_row_data = [cell.value for cell in self.sheet[2]]

        # 修改10001的值为其它 可变更复制的行数
        for row in range(3, 10001):
            for col, value in enumerate(first_row_data, start=1):
                self.sheet.cell(row=row, column=col, value=value)

        # 保存修改后的工作簿
        self.workbook.save(output_path)

    def increment_first_column(self, output_path):
        """
        从第二行开始，每一行第一列的值递增，其余的列复制第二行的内容
        :param output_path: 处理后输出文件路径
        """
        start_value = 8666461001

        # 获取第二行的数据（从第二列开始）
        second_row_data = [cell.value for cell in self.sheet[2]][1:]

        # 从第二行开始，每一行第一列的值递增，并复制第二行的其余列的内容
        for row in range(2, 10001):
            # 设置第一列的递增值
            self.sheet.cell(row=row, column=1, value=start_value)
            start_value += 1

            # 复制第二行的其余列内容
            for col, value in enumerate(second_row_data, start=2):
                self.sheet.cell(row=row, column=col, value=value)

        # 保存修改后的文件
        self.workbook.save(output_path)

    def modify_cell(self, row, column, value):
        # 修改当前文件指定单元格的内容 row:行 column:列 value:值
        self.sheet.cell(row=row, column=column, value=value)
        self.workbook.save(self.file_path)

    def pd_read_excel(self):
        # 读取excel文件
        df = pd.read_excel(self.file_path).head()
        return df


if __name__ == '__main__':
    ...
    # ExcelModifier('/Users/wl/Downloads/case备份/big_new_file.xlsx').modify_cell(row=2, column=1, value='666')
    # ExcelModifier('/Users/wl/Downloads/case备份/big_new_file.xlsx').duplicate_first_row(output_path='/Users/wl/Downloads/case备份/big_new_file_1.xlsx')
    # print(ExcelModifier('/Users/wl/Downloads/case备份/big_new_file.xlsx').pd_read_excel())